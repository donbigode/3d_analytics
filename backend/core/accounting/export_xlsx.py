import io
from datetime import date
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.accounting.dre import compute_dre
from backend.core.accounting.facts import compute_facts
from backend.core.accounting.monthly import compute_dre_monthly
from backend.core.accounting.profitability import compute_profitability
from backend.core.models import ExpenseCategory
from backend.infra.db.models import Client, Expense, MaterialConsumption, QuoteItem, Sale

_CAT_LABEL = {"maintenance": "Manutenção", "parts": "Peças", "tools": "Ferramentas",
              "labor": "Mecânicos", "equipment": "Máquinas/Equipamentos", "other": "Outros"}


async def build_dre_xlsx(session: AsyncSession, period_from: date, period_to: date) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "DRE mensal"
    monthly = await compute_dre_monthly(session, period_from, period_to)
    total = await compute_dre(session, period_from, period_to)
    months = [m["month"] for m in monthly]
    ws.append(["Conta", *months, "Total"])

    def line(label, key):
        ws.append([label, *[float(m[key]) for m in monthly], float(total[key])])

    line("Receita bruta", "receita_bruta")
    line("(-) Impostos", "impostos")
    line("= Receita líquida", "receita_liquida")
    line("(-) CPV", "cpv")
    line("(-) Custos variáveis", "custos_variaveis")
    line("= Lucro bruto", "lucro_bruto")
    for cat in ExpenseCategory:
        ws.append([f"(-) {_CAT_LABEL[cat.value]}",
                   *[float(m["despesas"][cat.value]) for m in monthly],
                   float(total["despesas"][cat.value])])
    line("(-) Custo de estoque", "custo_estoque")
    line("= Resultado líquido", "resultado_liquido")
    line("Margem líquida %", "margem_liquida_pct")

    ws = wb.create_sheet("Vendas")
    ws.append(["Orçamento", "Cliente", "Tipo", "Status", "Total", "CPV", "Receita", "Variáveis", "Data"])
    sales = (await session.execute(select(Sale).where(
        Sale.is_sold.is_(True), Sale.is_stale.is_(False),
        Sale.sold_at >= period_from, Sale.sold_at <= period_to))).scalars().all()
    for s in sales:
        cname = ""
        if s.client_id:
            c = await session.get(Client, s.client_id)
            cname = c.name if c else ""
        ws.append([str(s.quote_id), cname, s.quote_kind, s.quote_status, float(s.quote_total),
                   float(s.cpv_calc), float(s.confirmed_revenue or 0), float(s.variable_costs),
                   s.sold_at.isoformat() if s.sold_at else ""])

    ws = wb.create_sheet("Despesas")
    ws.append(["Categoria", "Descrição", "Valor", "Recorrente", "Data"])
    exps = (await session.execute(select(Expense).where(
        Expense.incurred_at >= period_from, Expense.incurred_at <= period_to))).scalars().all()
    for e in exps:
        ws.append([_CAT_LABEL.get(e.category, e.category), e.description, float(e.amount),
                   "sim" if e.is_recurring else "não", e.incurred_at.isoformat()])

    ws = wb.create_sheet("Custo de estoque")
    ws.append(["Orçamento", "Gramas", "Custo unit.", "Total", "Data"])
    rows = (await session.execute(
        select(MaterialConsumption, QuoteItem.quote_id)
        .join(QuoteItem, MaterialConsumption.quote_item_id == QuoteItem.id)
        .where(MaterialConsumption.consumed_at >= period_from))).all()
    for cons, qid in rows:
        ws.append([str(qid), float(cons.grams_used), float(cons.unit_cost_snapshot),
                   float(cons.grams_used * cons.unit_cost_snapshot),
                   cons.consumed_at.date().isoformat() if cons.consumed_at else ""])

    ws = wb.create_sheet("Lucratividade")
    prof = await compute_profitability(session, period_from, period_to)
    ws.append(["Por cliente", "Receita", "Custo", "Margem", "Margem %"])
    for r in prof["by_client"]:
        ws.append([r["label"], float(r["receita"]), float(r["custo"]), float(r["margem"]), float(r["margem_pct"])])
    ws.append([])
    ws.append(["Por material", "Receita", "Custo", "Margem", "Margem %"])
    for r in prof["by_material"]:
        ws.append([r["label"], float(r["receita"]), float(r["custo"]), float(r["margem"]), float(r["margem_pct"])])

    ws = wb.create_sheet("Fato (itens)")
    cols = ["sale_id", "quote_id", "quote_kind", "cliente", "status", "sold_at", "is_sold",
            "receita_venda", "custos_variaveis_venda", "cpv_venda", "item_id", "nome",
            "quantidade", "material_type", "cor_material", "cor_bobina", "filament_m",
            "filament_g", "gramas_total", "custo_filamento_item", "receita_item"]
    ws.append(cols)
    for row in await compute_facts(session, period_from, period_to):
        ws.append([
            row["sale_id"], row["quote_id"], row["quote_kind"], row["cliente"], row["status"],
            row["sold_at"].isoformat() if row["sold_at"] else "", row["is_sold"],
            float(row["receita_venda"]), float(row["custos_variaveis_venda"]), float(row["cpv_venda"]),
            row["item_id"], row["nome"], row["quantidade"], row["material_type"],
            row["cor_material"] or "", row["cor_bobina"] or "",
            row["filament_m"] if row["filament_m"] is not None else "",
            row["filament_g"] if row["filament_g"] is not None else "",
            float(row["gramas_total"]), float(row["custo_filamento_item"]), float(row["receita_item"]),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
